import logging
import os
import sys

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook


# ============================================================
# CONFIGURATION
# ============================================================

EXCEL_FILE = "Birthdays.xlsx"
SOURCE_WORKSHEET = "Birthday_Calendar"

CREDENTIALS_FILE = "credentials.json"
DESTINATION_SPREADSHEET_ID = os.getenv("BIRTHDAY_CALENDAR_SHEET_ID")
DESTINATION_WORKSHEET = "Birthday Calendar"


# ============================================================
# LOGGING
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)

logger = logging.getLogger(__name__)


# ============================================================
# READ CLEAN CALENDAR FROM EXCEL
# ============================================================

def read_calendar_from_excel():
    """
    Read ONLY Month, Date and Name from the internal Excel
    Birthday_Calendar sheet.

    No Raw_Data, timestamps, anniversaries or duplicate-audit
    information is exposed to the public/shareable sheet.
    """

    logger.info("Reading calendar from %s...", EXCEL_FILE)

    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(
            f"Excel file not found: {EXCEL_FILE}"
        )

    workbook = load_workbook(
        EXCEL_FILE,
        data_only=True,
        read_only=True,
    )

    if SOURCE_WORKSHEET not in workbook.sheetnames:
        workbook.close()
        raise ValueError(
            f"Worksheet '{SOURCE_WORKSHEET}' not found in {EXCEL_FILE}"
        )

    worksheet = workbook[SOURCE_WORKSHEET]

    header_row = [
        cell.value
        for cell in worksheet[1]
    ]

    required_columns = [
        "Month",
        "Date",
        "Name",
    ]

    missing = [
        column
        for column in required_columns
        if column not in header_row
    ]

    if missing:
        workbook.close()
        raise ValueError(
            f"Missing required Birthday_Calendar columns: {missing}"
        )

    indexes = {
        column: header_row.index(column)
        for column in required_columns
    }

    output_rows = [
        required_columns
    ]

    for row in worksheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        month = row[indexes["Month"]]
        day = row[indexes["Date"]]
        name = row[indexes["Name"]]

        if month is None and day is None and name is None:
            continue

        output_rows.append(
            [
                "" if month is None else str(month),
                "" if day is None else day,
                "" if name is None else str(name).strip(),
            ]
        )

    workbook.close()

    logger.info(
        "Prepared %s birthday calendar rows for publishing.",
        len(output_rows) - 1,
    )

    return output_rows


# ============================================================
# GOOGLE SHEETS CONNECTION
# ============================================================

def connect_to_destination_sheet():
    """
    Connect to the separate shareable Google Spreadsheet.

    The service-account email from credentials.json must be
    shared on that spreadsheet with Editor access.
    """

    if not DESTINATION_SPREADSHEET_ID:
        raise ValueError(
            "BIRTHDAY_CALENDAR_SHEET_ID is not set."
        )

    if not os.path.exists(CREDENTIALS_FILE):
        raise FileNotFoundError(
            f"Credentials file not found: {CREDENTIALS_FILE}"
        )

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
    ]

    credentials = Credentials.from_service_account_file(
        CREDENTIALS_FILE,
        scopes=scopes,
    )

    client = gspread.authorize(credentials)

    spreadsheet = client.open_by_key(
        DESTINATION_SPREADSHEET_ID
    )

    try:
        worksheet = spreadsheet.worksheet(
            DESTINATION_WORKSHEET
        )

    except gspread.WorksheetNotFound:
        worksheet = spreadsheet.add_worksheet(
            title=DESTINATION_WORKSHEET,
            rows=200,
            cols=3,
        )

    return worksheet


# ============================================================
# PUBLISH
# ============================================================

def publish_calendar(rows):
    """
    Replace the shareable calendar sheet with the latest
    Birthday_Calendar output.
    """

    worksheet = connect_to_destination_sheet()

    logger.info(
        "Publishing latest calendar to Google Sheets..."
    )

    # Clear previous published output so removed/changed rows
    # do not remain visible.
    worksheet.clear()

    worksheet.update(
        range_name="A1",
        values=rows,
        value_input_option="RAW",
    )

    # Small presentation improvements.
    try:
        worksheet.freeze(rows=1)

        worksheet.format(
            "A1:C1",
            {
                "textFormat": {
                    "bold": True,
                },
                "horizontalAlignment": "CENTER",
            },
        )
    except Exception as formatting_error:
        # Formatting should never prevent the data itself from
        # being published.
        logger.warning(
            "Calendar published, but formatting could not "
            "be fully applied: %s",
            formatting_error,
        )

    logger.info(
        "Published %s birthday rows successfully.",
        len(rows) - 1,
    )


# ============================================================
# MAIN
# ============================================================

def main():
    logger.info(
        "========== Birthday Calendar Publisher Started =========="
    )

    try:
        rows = read_calendar_from_excel()
        publish_calendar(rows)

        logger.info(
            "========== Birthday Calendar Published Successfully =========="
        )

    except Exception as error:
        logger.exception(
            "PUBLISH FAILED: %s",
            error,
        )
        sys.exit(1)


if __name__ == "__main__":
    main()
