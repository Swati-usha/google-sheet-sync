import json
import logging
import os
import sys
from collections import defaultdict
from datetime import datetime

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# CONFIGURATION
# ============================================================

GOOGLE_SHEET_NAME = "Birthday Memo"
GOOGLE_WORKSHEET_NAME = "Form Responses 1"

CREDENTIALS_FILE = "credentials.json"
WATERMARK_FILE = "watermark.json"
EXCEL_FILE = "Birthdays.xlsx"

RAW_SHEET = "Raw_Data"
BIRTHDAY_VIEW_SHEET = "Birthday_View"
BIRTHDAY_CALENDAR_SHEET = "Birthday_Calendar"
DUPLICATE_SHEET = "Duplicate_Check"

LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()


# ============================================================
# LOGGING
# ============================================================

logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s | %(levelname)s | %(message)s",
)

logger = logging.getLogger(__name__)


# ============================================================
# GOOGLE SHEETS CONNECTION
# ============================================================

def connect_to_google_sheet():
    """Connect to Google Sheets using service-account credentials."""

    logger.info("Connecting to Google Sheets...")

    if not os.path.exists(CREDENTIALS_FILE):
        raise FileNotFoundError(
            f"Credentials file not found: {CREDENTIALS_FILE}"
        )

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]

    credentials = Credentials.from_service_account_file(
        CREDENTIALS_FILE,
        scopes=scopes,
    )

    client = gspread.authorize(credentials)

    spreadsheet = client.open(GOOGLE_SHEET_NAME)

    worksheet = spreadsheet.worksheet(
        GOOGLE_WORKSHEET_NAME
    )

    logger.info(
        "Connected to '%s' → '%s'",
        GOOGLE_SHEET_NAME,
        GOOGLE_WORKSHEET_NAME,
    )

    return worksheet


# ============================================================
# WATERMARK
# ============================================================

def read_watermark():
    """Read the last successfully processed Google Sheet row."""

    if not os.path.exists(WATERMARK_FILE):

        logger.warning(
            "%s not found. Starting from row 1.",
            WATERMARK_FILE,
        )

        return 1

    try:

        with open(WATERMARK_FILE, "r") as file:
            data = json.load(file)

        last_row = int(
            data.get("last_row", 1)
        )

        if last_row < 1:
            raise ValueError(
                "Watermark must be >= 1."
            )

        return last_row

    except (
        json.JSONDecodeError,
        ValueError,
        TypeError,
    ) as error:

        raise ValueError(
            f"Invalid watermark file: {error}"
        )


def write_watermark(last_row):
    """
    Persist watermark only after Excel update succeeds.
    """

    temporary_file = f"{WATERMARK_FILE}.tmp"

    with open(
        temporary_file,
        "w",
    ) as file:

        json.dump(
            {"last_row": last_row},
            file,
            indent=4,
        )

    os.replace(
        temporary_file,
        WATERMARK_FILE,
    )

    logger.info(
        "Watermark updated to row %s",
        last_row,
    )


# ============================================================
# EXTRACT
# ============================================================

def get_new_rows(worksheet, last_row):
    """
    Read Google Sheet and return rows after watermark.

    Google Sheet:
        Row 1 = header
        Row 2+ = data
    """

    logger.info(
        "Last processed row: %s",
        last_row,
    )

    values = worksheet.get_all_values()

    actual_last_row = len(values)

    logger.info(
        "Actual Google Sheet rows: %s",
        actual_last_row,
    )

    if last_row > actual_last_row:

        raise ValueError(
            f"Watermark ({last_row}) is ahead of "
            f"Google Sheet ({actual_last_row}). "
            "Manual intervention required."
        )

    if actual_last_row == last_row:

        logger.info(
            "No new rows found."
        )

        return [], last_row

    new_rows = values[last_row:]

    new_last_row = actual_last_row

    logger.info(
        "Fetched %s new rows.",
        len(new_rows),
    )

    logger.info(
        "New last row: %s",
        new_last_row,
    )

    return new_rows, new_last_row


# ============================================================
# EXCEL - RAW DATA
# ============================================================

def create_excel(headers):
    """Create a new Excel workbook with Raw_Data."""

    logger.info(
        "Creating new Excel file: %s",
        EXCEL_FILE,
    )

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = RAW_SHEET

    worksheet.append(headers)

    workbook.save(EXCEL_FILE)

    return workbook, worksheet


def append_to_excel(headers, new_rows):
    """Append new Google Sheet rows to Raw_Data."""

    if not new_rows:
        return

    if not os.path.exists(EXCEL_FILE):

        workbook, worksheet = create_excel(
            headers
        )

    else:

        logger.info(
            "Opening existing Excel file: %s",
            EXCEL_FILE,
        )

        workbook = load_workbook(
            EXCEL_FILE
        )

        if RAW_SHEET in workbook.sheetnames:

            worksheet = workbook[RAW_SHEET]

        elif "Responses" in workbook.sheetnames:

            logger.info(
                "Renaming existing 'Responses' sheet to '%s'.",
                RAW_SHEET,
            )

            worksheet = workbook["Responses"]

            worksheet.title = RAW_SHEET

        else:

            worksheet = workbook.create_sheet(
                RAW_SHEET
            )

            worksheet.append(headers)

    for row in new_rows:

        worksheet.append(row)

    workbook.save(EXCEL_FILE)

    workbook.close()

    logger.info(
        "Appended %s rows to Excel.",
        len(new_rows),
    )


# ============================================================
# DATE / NAME HELPERS
# ============================================================

def parse_birthday(value):
    """
    Extract month and day from birthday value.

    Supported examples:

        8/4/1987
        12/29/1986
        7/21/1988
    """

    if not value:
        return None

    value = str(value).strip()

    formats = [
        "%m/%d/%Y",
        "%m/%d/%y",
    ]

    for fmt in formats:

        try:

            date_value = datetime.strptime(
                value,
                fmt,
            )

            return (
                date_value.month,
                date_value.day,
            )

        except ValueError:
            continue

    return None


def normalize_dob(value):
    """Convert DOB into a comparable date."""

    if not value:
        return None

    if isinstance(value, datetime):

        return value.date()

    value = str(value).strip()

    formats = [
        "%m/%d/%Y",
        "%m/%d/%y",
    ]

    for fmt in formats:

        try:

            return datetime.strptime(
                value,
                fmt,
            ).date()

        except ValueError:
            continue

    return None


def normalize_name(name):
    """
    Normalize name for comparison.

    Example:

        "  Swati   Tyagi "
            →
        "swati tyagi"
    """

    if not name:
        return ""

    return " ".join(
        str(name)
        .strip()
        .lower()
        .split()
    )


def get_first_name(name):
    """Return normalized first name."""

    normalized = normalize_name(name)

    if not normalized:
        return ""

    return normalized.split()[0]


def is_first_name_only(name):
    """Return True when name contains exactly one word."""

    normalized = normalize_name(name)

    if not normalized:
        return False

    return len(
        normalized.split()
    ) == 1


# ============================================================
# DUPLICATE RESOLUTION
# ============================================================

def resolve_duplicate_records(records):
    """
    Resolve duplicate birthday records safely.

    Duplicate rule:

    A first-name-only record and a fuller-name record
    are considered duplicates ONLY when:

        1. First name is the same
        2. DOB is the same

    Example:

        Swati       | 10/1/1987
        Swati Tyagi | 10/1/1987

    Result:

        Keep Swati Tyagi

    But:

        Rahul Sharma | 1/5/1988
        Rahul Verma  | 1/5/1988

    are NOT considered duplicates because both are
    already full names.

    If multiple first-name-only records exist with the
    same DOB, only the first is retained.
    """

    groups = defaultdict(list)

    # --------------------------------------------------------
    # Group by first name + DOB
    # --------------------------------------------------------

    for record in records:

        first_name = get_first_name(
            record["name"]
        )

        dob = normalize_dob(
            record["birthday"]
        )

        if not first_name or not dob:
            continue

        key = (
            first_name,
            dob,
        )

        groups[key].append(record)

    resolved_records = []

    duplicate_records = []

    grouped_ids = set()

    # --------------------------------------------------------
    # Resolve groups
    # --------------------------------------------------------

    for group in groups.values():

        for record in group:

            grouped_ids.add(
                id(record)
            )

        first_name_only_records = [
            record
            for record in group
            if is_first_name_only(
                record["name"]
            )
        ]

        full_name_records = [
            record
            for record in group
            if not is_first_name_only(
                record["name"]
            )
        ]

        # ----------------------------------------------------
        # CASE 1:
        # There is a full-name record.
        #
        # Keep full-name record.
        # Ignore first-name-only records.
        # ----------------------------------------------------

        if full_name_records:

            # If there is exactly one full name,
            # keep it.

            if len(full_name_records) == 1:

                best_record = (
                    full_name_records[0]
                )

            else:

                # Multiple full names with same
                # first name + DOB are ambiguous.
                #
                # Keep the longest one for the
                # calendar, but record the others.

                best_record = max(
                    full_name_records,
                    key=lambda record: len(
                        normalize_name(
                            record["name"]
                        )
                    ),
                )

            resolved_records.append(
                best_record
            )

            # First-name-only duplicates

            for record in first_name_only_records:

                duplicate_records.append(
                    {
                        "Kept_Name": best_record["name"],
                        "Kept_Birthday": best_record["birthday"],
                        "Ignored_Name": record["name"],
                        "Ignored_Birthday": record["birthday"],
                        "Reason": (
                            "Same first name and DOB; "
                            "full-name entry kept"
                        ),
                    }
                )

            # Additional full-name records

            for record in full_name_records:

                if record is not best_record:

                    duplicate_records.append(
                        {
                            "Kept_Name": best_record["name"],
                            "Kept_Birthday": best_record["birthday"],
                            "Ignored_Name": record["name"],
                            "Ignored_Birthday": record["birthday"],
                            "Reason": (
                                "Same first name and DOB; "
                                "multiple full-name entries"
                            ),
                        }
                    )

        # ----------------------------------------------------
        # CASE 2:
        # Only first-name records exist.
        # ----------------------------------------------------

        elif first_name_only_records:

            best_record = (
                first_name_only_records[0]
            )

            resolved_records.append(
                best_record
            )

            for record in (
                first_name_only_records[1:]
            ):

                duplicate_records.append(
                    {
                        "Kept_Name": best_record["name"],
                        "Kept_Birthday": best_record["birthday"],
                        "Ignored_Name": record["name"],
                        "Ignored_Birthday": record["birthday"],
                        "Reason": (
                            "Same first name and DOB; "
                            "duplicate first-name entry"
                        ),
                    }
                )

    # --------------------------------------------------------
    # Records that cannot participate in duplicate matching
    # --------------------------------------------------------

    for record in records:

        if id(record) not in grouped_ids:

            resolved_records.append(
                record
            )

    return (
        resolved_records,
        duplicate_records,
    )


# ============================================================
# BIRTHDAY CURRENT
# ============================================================

def is_birthday_current(month, day):
    """
    Return TRUE only when birthday is today.

    Example:

        Today = August 13

        August 13 -> TRUE
        August 14 -> FALSE
        August 12 -> FALSE
    """

    if not month or not day:
        return False

    today = datetime.today().date()

    return (
        today.month == month
        and today.day == day
    )


# ============================================================
# EXCEL FORMATTING
# ============================================================

def auto_size_columns(
    worksheet,
    max_width=30,
):
    """Automatically size worksheet columns."""

    for column in worksheet.columns:

        max_length = 0

        column_letter = (
            column[0].column_letter
        )

        for cell in column:

            if cell.value is not None:

                max_length = max(
                    max_length,
                    len(str(cell.value)),
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            max_width,
        )


# ============================================================
# BIRTHDAY VIEW
# ============================================================

def refresh_birthday_view(
    workbook,
    resolved_records,
):
    """
    Refresh Birthday_View.

    This provides the detailed birthday data after
    duplicate resolution.
    """

    logger.info(
        "Refreshing Birthday_View..."
    )

    if BIRTHDAY_VIEW_SHEET in workbook.sheetnames:

        del workbook[
            BIRTHDAY_VIEW_SHEET
        ]

    birthday_sheet = workbook.create_sheet(
        BIRTHDAY_VIEW_SHEET
    )

    birthday_sheet.append(
        [
            "Name",
            "Birthday",
            "Month",
            "Day",
            "Birthday_Current",
        ]
    )

    # --------------------------------------------------------
    # Sort records
    # --------------------------------------------------------

    resolved_records.sort(
        key=lambda record: (
            record["month"],
            record["day"],
            normalize_name(
                record["name"]
            ),
        )
    )

    # --------------------------------------------------------
    # Write records
    # --------------------------------------------------------

    for record in resolved_records:

        month = record["month"]
        day = record["day"]

        month_name = datetime(
            2000,
            month,
            1,
        ).strftime("%B")

        birthday_current = (
            is_birthday_current(
                month,
                day,
            )
        )

        birthday_sheet.append(
            [
                record["name"],
                record["birthday"],
                month_name,
                day,
                birthday_current,
            ]
        )

    # --------------------------------------------------------
    # Formatting
    # --------------------------------------------------------

    birthday_sheet.freeze_panes = "A2"

    birthday_sheet.auto_filter.ref = (
        birthday_sheet.dimensions
    )

    for cell in birthday_sheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    # --------------------------------------------------------
    # Highlight today's birthdays
    # --------------------------------------------------------

    current_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    if birthday_sheet.max_row >= 2:

        birthday_sheet.conditional_formatting.add(
            f"A2:E{birthday_sheet.max_row}",
            FormulaRule(
                formula=[
                    "$E2=TRUE"
                ],
                fill=current_fill,
            ),
        )

    auto_size_columns(
        birthday_sheet,
        max_width=30,
    )

    logger.info(
        "Birthday_View refreshed with %s records.",
        len(resolved_records),
    )


# ============================================================
# BIRTHDAY CALENDAR
# ============================================================

def refresh_birthday_calendar(
    workbook,
    resolved_records,
):
    """
    Create the final Birthday_Calendar view.

    Only dates that actually contain birthdays are shown.

    Example:

        January
            5   Sanjeev Kumar
           22   Shrey Dhingra

        August
            4   Shatakshi Verma
            7   Harsh Dutt
            8   Vikrant

    Records are sorted chronologically.
    """

    logger.info(
        "Refreshing Birthday_Calendar..."
    )

    if BIRTHDAY_CALENDAR_SHEET in workbook.sheetnames:

        del workbook[
            BIRTHDAY_CALENDAR_SHEET
        ]

    calendar_sheet = workbook.create_sheet(
        BIRTHDAY_CALENDAR_SHEET
    )

    calendar_sheet.append(
        [
            "Month",
            "Date",
            "Name",
            "Birthday",
            "Birthday_Current",
        ]
    )

    # --------------------------------------------------------
    # Sort
    # --------------------------------------------------------

    sorted_records = sorted(
        resolved_records,
        key=lambda record: (
            record["month"],
            record["day"],
            normalize_name(
                record["name"]
            ),
        ),
    )

    dates_created = set()

    # --------------------------------------------------------
    # Write calendar
    # --------------------------------------------------------

    for record in sorted_records:

        month = record["month"]
        day = record["day"]

        month_name = datetime(
            2000,
            month,
            1,
        ).strftime("%B")

        birthday_current = (
            is_birthday_current(
                month,
                day,
            )
        )

        calendar_sheet.append(
            [
                month_name,
                day,
                record["name"],
                record["birthday"],
                birthday_current,
            ]
        )

        dates_created.add(
            (
                month,
                day,
            )
        )

    # --------------------------------------------------------
    # Formatting
    # --------------------------------------------------------

    calendar_sheet.freeze_panes = "A2"

    calendar_sheet.auto_filter.ref = (
        calendar_sheet.dimensions
    )

    for cell in calendar_sheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    # --------------------------------------------------------
    # Highlight today's birthdays
    # --------------------------------------------------------

    current_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    if calendar_sheet.max_row >= 2:

        calendar_sheet.conditional_formatting.add(
            f"A2:E{calendar_sheet.max_row}",
            FormulaRule(
                formula=[
                    "$E2=TRUE"
                ],
                fill=current_fill,
            ),
        )

    auto_size_columns(
        calendar_sheet,
        max_width=30,
    )

    logger.info(
        "Birthday_Calendar refreshed with %s dates.",
        len(dates_created),
    )


# ============================================================
# DUPLICATE CHECK
# ============================================================

def refresh_duplicate_check(
    workbook,
    duplicate_records,
):
    """
    Rebuild Duplicate_Check.

    This sheet shows which records were ignored by
    duplicate resolution and why.
    """

    logger.info(
        "Refreshing Duplicate_Check..."
    )

    if DUPLICATE_SHEET in workbook.sheetnames:

        del workbook[
            DUPLICATE_SHEET
        ]

    duplicate_sheet = workbook.create_sheet(
        DUPLICATE_SHEET
    )

    duplicate_sheet.append(
        [
            "Kept_Name",
            "Kept_Birthday",
            "Ignored_Name",
            "Ignored_Birthday",
            "Reason",
        ]
    )

    for cell in duplicate_sheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    for duplicate in duplicate_records:

        duplicate_sheet.append(
            [
                duplicate["Kept_Name"],
                duplicate["Kept_Birthday"],
                duplicate["Ignored_Name"],
                duplicate["Ignored_Birthday"],
                duplicate["Reason"],
            ]
        )

    duplicate_sheet.freeze_panes = "A2"

    duplicate_sheet.auto_filter.ref = (
        duplicate_sheet.dimensions
    )

    auto_size_columns(
        duplicate_sheet,
        max_width=50,
    )

    logger.info(
        "Duplicate_Check contains %s records.",
        len(duplicate_records),
    )


# ============================================================
# BUILD DERIVED VIEWS
# ============================================================

def refresh_all_views():
    """
    Rebuild all derived Excel sheets from Raw_Data.

    This runs every time sync.py runs, even when there
    are no new Google Sheet records.

    This is important because Birthday_Current depends
    on today's date.
    """

    logger.info(
        "Refreshing all derived birthday views..."
    )

    if not os.path.exists(EXCEL_FILE):

        logger.warning(
            "Excel file does not exist. "
            "Cannot refresh views."
        )

        return

    workbook = load_workbook(
        EXCEL_FILE
    )

    # --------------------------------------------------------
    # Locate Raw_Data
    # --------------------------------------------------------

    if RAW_SHEET in workbook.sheetnames:

        raw_sheet = workbook[
            RAW_SHEET
        ]

    elif "Responses" in workbook.sheetnames:

        logger.info(
            "Renaming existing 'Responses' sheet to '%s'.",
            RAW_SHEET,
        )

        raw_sheet = workbook[
            "Responses"
        ]

        raw_sheet.title = RAW_SHEET

    else:

        logger.warning(
            "Raw_Data sheet not found."
        )

        workbook.close()

        return

    # --------------------------------------------------------
    # Read headers
    # --------------------------------------------------------

    headers = [
        cell.value
        for cell in raw_sheet[1]
    ]

    try:

        birthday_column = headers.index(
            "Mention your birthday"
        )

        name_column = headers.index(
            "Name"
        )

    except ValueError as error:

        workbook.close()

        raise ValueError(
            f"Required Raw_Data column missing: {error}"
        )

    # --------------------------------------------------------
    # Read records
    # --------------------------------------------------------

    records = []

    for row in raw_sheet.iter_rows(
        min_row=2,
        values_only=True,
    ):

        name = row[name_column]

        birthday = row[
            birthday_column
        ]

        if not name and not birthday:
            continue

        parsed = parse_birthday(
            birthday
        )

        if not parsed:

            logger.warning(
                "Skipping invalid birthday: %s | Name: %s",
                birthday,
                name,
            )

            continue

        month, day = parsed

        records.append(
            {
                "name": str(name).strip()
                if name
                else "",

                "birthday": str(
                    birthday
                ).strip()
                if birthday
                else "",

                "month": month,

                "day": day,
            }
        )

    logger.info(
        "Read %s records from Raw_Data.",
        len(records),
    )

    # --------------------------------------------------------
    # Resolve duplicates
    # --------------------------------------------------------

    resolved_records, duplicate_records = (
        resolve_duplicate_records(
            records
        )
    )

    logger.info(
        "Records after duplicate resolution: %s",
        len(resolved_records),
    )

    logger.info(
        "Duplicate records detected: %s",
        len(duplicate_records),
    )

    # --------------------------------------------------------
    # Refresh views
    # --------------------------------------------------------

    refresh_birthday_view(
        workbook,
        resolved_records,
    )

    refresh_birthday_calendar(
        workbook,
        resolved_records,
    )

    refresh_duplicate_check(
        workbook,
        duplicate_records,
    )

    # --------------------------------------------------------
    # Save
    # --------------------------------------------------------

    workbook.save(
        EXCEL_FILE
    )

    workbook.close()

    logger.info(
        "All derived views refreshed successfully."
    )


# ============================================================
# MAIN
# ============================================================

def main():

    logger.info(
        "========== Birthday Memo Sync Started =========="
    )

    try:

        # ----------------------------------------------------
        # 1. Read watermark
        # ----------------------------------------------------

        last_row = read_watermark()

        # ----------------------------------------------------
        # 2. Connect to Google Sheets
        # ----------------------------------------------------

        worksheet = connect_to_google_sheet()

        # ----------------------------------------------------
        # 3. Read headers
        # ----------------------------------------------------

        all_values = worksheet.get_all_values()

        if not all_values:

            raise ValueError(
                "Google Sheet is empty."
            )

        headers = all_values[0]

        logger.info(
            "Detected %s columns.",
            len(headers),
        )

        # ----------------------------------------------------
        # 4. Extract new rows
        # ----------------------------------------------------

        new_rows, new_last_row = (
            get_new_rows(
                worksheet,
                last_row,
            )
        )

        # ----------------------------------------------------
        # 5. Incremental load
        # ----------------------------------------------------

        if new_rows:

            append_to_excel(
                headers,
                new_rows,
            )

            # ------------------------------------------------
            # Update watermark only after Excel succeeds
            # ------------------------------------------------

            write_watermark(
                new_last_row
            )

        else:

            logger.info(
                "No new rows to append."
            )

        # ----------------------------------------------------
        # 6. Refresh all derived views
        # ----------------------------------------------------

        refresh_all_views()

        # ----------------------------------------------------
        # 7. Completion
        # ----------------------------------------------------

        if new_rows:

            logger.info(
                "========== Sync Completed Successfully =========="
            )

        else:

            logger.info(
                "========== Sync Completed: "
                "Views Refreshed, Nothing New to Load =========="
            )

    except Exception as error:

        logger.exception(
            "SYNC FAILED: %s",
            error,
        )

        # Important for GitHub Actions
        sys.exit(1)


# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":
    main()